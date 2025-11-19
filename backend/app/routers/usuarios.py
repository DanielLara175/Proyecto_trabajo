from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from sqlalchemy.orm import Session
from typing import List
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook

# Importación correcta de schemas y crud
from app import crud, schemas, models
from app.database import get_db

# Crear router para las rutas relacionadas con usuarios
router = APIRouter(
    prefix="/usuarios",
    tags=["Usuarios"]
)

# Crear un nuevo usuario
@router.post("/", response_model=schemas.UsuarioRead, status_code=status.HTTP_201_CREATED)
def crear_usuario(usuario: schemas.UsuarioCreate, db: Session = Depends(get_db)):
    """
    Crea un nuevo usuario en la base de datos.
    """
    return crud.crear_usuario(db, usuario)


# Listar todos los usuarios
@router.get("/", response_model=List[schemas.UsuarioRead])
def listar_usuarios(db: Session = Depends(get_db)):
    """
    Retorna una lista de todos los usuarios registrados.
    """
    return crud.obtener_usuarios(db)


# Eliminar un usuario por ID
@router.delete("/{usuario_id}", status_code=status.HTTP_200_OK)
def eliminar_usuario(usuario_id: int, db: Session = Depends(get_db)):
    """
    Elimina un usuario existente por su ID.
    """
    return crud.borrar_usuario(db, usuario_id)


# -----------------------------------------------------------
# Endpoint para importar usuarios desde un archivo Excel (.xlsx)
# -----------------------------------------------------------

@router.post("/importar-excel")
async def importar_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """
    Importa usuarios desde un archivo Excel.
    El archivo debe tener las columnas 'name' y 'email'.
    Cada fila será registrada en la base de datos como un nuevo usuario.
    """

    # Validación de extensión del archivo
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(
            status_code=400,
            detail="El archivo debe ser formato .xlsx"
        )

    try:
        # Leer archivo completamente en memoria
        contenido = file.file.read()

        # Convertir contenido binario a un buffer para openpyxl
        buffer = BytesIO(contenido)

        # Cargar Excel usando openpyxl
        workbook = load_workbook(buffer)
        sheet = workbook.active

        # Verificar encabezados correctos
        encabezados = [cell.value for cell in sheet[1]]
        if "name" not in encabezados or "email" not in encabezados:
            raise HTTPException(
                status_code=400,
                detail="El archivo debe contener columnas 'name' y 'email'"
            )

        # Obtener índices de columnas
        idx_name = encabezados.index("name")
        idx_email = encabezados.index("email")

        # Recorrer filas desde la fila 2 hacia abajo
        for fila in sheet.iter_rows(min_row=2, values_only=True):
            nombre = fila[idx_name]
            email = fila[idx_email]

            # Ignorar filas vacías
            if not nombre or not email:
                continue

            # Crear instancia del modelo
            nuevo_usuario = models.User(
                name=nombre,
                email=email
            )
            db.add(nuevo_usuario)

        # Confirmar cambios
        db.commit()

        return {"mensaje": "Usuarios importados correctamente"}

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error al procesar el archivo: {str(e)}"
        )
